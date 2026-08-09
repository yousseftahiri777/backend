"""Build COD network Excel exports from the official leads template."""

from __future__ import annotations

import io
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from app.models import Order
from app.phone_utils import format_ksa_phone_international
from app.product_catalog import (
    get_cod_network_sku,
    get_export_product_name,
    get_product_url,
)

TEMPLATE_PATH = (
    Path(__file__).resolve().parent.parent.parent / "assets" / "cod_network_leads_template.xlsx"
)

# Exact headers from leads-template (1).xlsx
COD_NETWORK_HEADERS = [
    "OrderDate",
    "country",
    "name",
    "phone",
    "address",
    "url",
    "sku",
    "Product",
    "quantity",
    "price",
    "currency",
    "notes",
    "utm_source",
    "utm_medium",
    "utm_campaign",
    "utm_term",
    "utm_content",
    "national_address",
]

# Template sample uses "SAUDIA " with a trailing space.
COD_NETWORK_COUNTRY = "SAUDIA "


def _order_date_value(created_at: datetime) -> date:
    return created_at.date()


def _phone_value(phone: str) -> int | None:
    digits = format_ksa_phone_international(phone or "")
    if not digits:
        return None
    return int(digits)


def _line_price(order: Order, order_items: list[dict], idx: int, item: dict) -> float:
    qty = int(item.get("qty") or 1)
    if len(order_items) == 1:
        return float(order.total)
    unit = float(item.get("price") or 0)
    line_price = round(unit * qty, 2)
    if idx == 0 and order.shipping:
        line_price = round(line_price + float(order.shipping or 0), 2)
    return line_price


def build_cod_network_rows(orders: Iterable[Order]) -> list[list[object]]:
    rows: list[list[object]] = []
    for order in orders:
        order_items = order.items or []
        if not order_items:
            continue

        phone = _phone_value(order.phone or "")

        for idx, item in enumerate(order_items):
            product_id = str(item.get("productId") or "")
            qty = int(item.get("qty") or 1)
            rows.append(
                [
                    _order_date_value(order.created_at),
                    COD_NETWORK_COUNTRY,
                    order.customer_name or "",
                    phone,
                    order.city or "",
                    get_product_url(product_id),
                    get_cod_network_sku(product_id),
                    get_export_product_name(
                        product_id, str(item.get("nameAr") or product_id)
                    ),
                    qty,
                    _line_price(order, order_items, idx, item),
                    "SAR",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                ]
            )
    return rows


def build_cod_network_workbook(orders: Iterable[Order]) -> io.BytesIO:
    if not TEMPLATE_PATH.is_file():
        raise FileNotFoundError(f"COD network template not found: {TEMPLATE_PATH}")

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # Keep row 1 (styled headers). Remove sample data rows.
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for row in build_cod_network_rows(orders):
        ws.append(row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
