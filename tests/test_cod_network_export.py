from datetime import datetime
from types import SimpleNamespace

from openpyxl import load_workbook

from app.services.cod_network_export import (
    COD_NETWORK_COUNTRY,
    COD_NETWORK_HEADERS,
    build_cod_network_rows,
    build_cod_network_workbook,
)


def _sample_order(**overrides):
    base = {
        "customer_name": "عبدالمجيد البريكي",
        "phone": "0504584137",
        "city": "الرياض",
        "total": 199.0,
        "shipping": 0.0,
        "created_at": datetime(2026, 8, 9, 12, 0, 0),
        "items": [
            {
                "productId": "sweat-shield",
                "nameAr": "درع العرق اليومي",
                "qty": 3,
                "price": 89,
            }
        ],
    }
    base.update(overrides)
    return SimpleNamespace(**base)


def test_cod_network_headers_match_template():
    from app.services.cod_network_export import TEMPLATE_PATH

    wb = load_workbook(TEMPLATE_PATH, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert headers[: len(COD_NETWORK_HEADERS)] == COD_NETWORK_HEADERS


def test_build_cod_network_rows():
    rows = build_cod_network_rows([_sample_order()])
    assert len(rows) == 1
    row = rows[0]
    assert row[0].isoformat() == "2026-08-09"
    assert row[1] == COD_NETWORK_COUNTRY
    assert row[2] == "عبدالمجيد البريكي"
    assert row[3] == 966504584137
    assert row[4] == "الرياض"
    assert row[5] == "https://lamabeauty.shop/products/sweat-shield"
    assert row[6] == "MP-YU0SFE1SC0RH"
    assert "50 pair" in row[7]
    assert row[8] == 3
    assert row[9] == 199.0
    assert row[10] == "SAR"
    assert row[11:] == ["", "", "", "", "", "", ""]


def test_build_cod_network_workbook():
    workbook = build_cod_network_workbook([_sample_order()])
    wb = load_workbook(workbook)
    ws = wb.active
    assert ws.max_row == 2
    assert ws["B2"].value == COD_NETWORK_COUNTRY
    assert ws["D2"].value == 966504584137
    assert ws["G2"].value == "MP-YU0SFE1SC0RH"
    assert ws["I2"].value == 3
    assert ws["J2"].value == 199
